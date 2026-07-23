#!/usr/bin/env python3
"""
fiopass.py - Gera formulários FIOTEC (Passagens e Diárias) a partir de arquivo de entrada.

Uso: python fiopass.py <arquivo_entrada.xls>
"""

import sys
import os
from copy import copy
from html.parser import HTMLParser
from datetime import datetime
import openpyxl

TEMPLATE_FILENAME = 'template_fiotec.xlsx'
VERSION_FILENAME = 'VERSION'

# Colunas esperadas no arquivo de entrada, na ordem, conforme
# formulario_320_2026-07-23_112029.xls. Usado para detectar arquivos gerados
# por uma versão desatualizada do formulário antes de tentar interpretá-los.
EXPECTED_HEADER = [
    'CPF', 'Desc. Pub.', 'Informe o nome da atividade de avaliação',
    'Período do deslocamento', 'Local da atividade de avaliação',
    'Itens solicitados', 'Nome Completo', 'Data de Nascimento', 'CPF',
    'Cargo/Função', 'Documento de Identificação', 'Nome do Banco', 'Agência',
    'Dígito da agência', 'Conta Corrente', 'Dígito da conta', 'Poupança',
    'Comprovante bancário', 'Tipo de logística necessária',
] + [
    label
    for _ in range(10)
    for label in (
        'Tipo de Trecho', 'Origem', 'Destino', 'Tipo Localidade de Destino',
        'Data', 'Período', 'Tipo Deslocamento', 'Observações sobre o deslocamento',
    )
] + [
    'Justificativa para solicitação fora do prazo',
    'Observações', 'Declaração de Compromisso',
]


class InputFormatError(Exception):
    """Levantado quando o arquivo de entrada não tem as colunas esperadas."""


def get_resource_path(filename):
    """Resolve o caminho do arquivo tanto em execução normal quanto em bundle PyInstaller."""
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, filename)
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), filename)


def get_version():
    """Lê a versão (data da última atualização do parsing) do arquivo VERSION."""
    try:
        with open(get_resource_path(VERSION_FILENAME), encoding='utf-8') as f:
            return f.read().strip()
    except OSError:
        return '?'


class TableParser(HTMLParser):
    """Lê arquivo XLS no formato HTML e extrai as linhas da tabela."""

    def __init__(self):
        super().__init__()
        self.rows = []
        self._current_row = []
        self._current_cell = ''
        self._in_cell = False

    def handle_starttag(self, tag, attrs):
        if tag == 'tr':
            self._current_row = []
        elif tag in ('td', 'th'):
            self._in_cell = True
            self._current_cell = ''

    def handle_endtag(self, tag):
        if tag in ('td', 'th'):
            self._current_row.append(self._current_cell.strip())
            self._in_cell = False
        elif tag == 'tr' and self._current_row:
            self.rows.append(self._current_row)

    def handle_data(self, data):
        if self._in_cell:
            self._current_cell += data


class _CellCollector(HTMLParser):
    """Coleta o texto de todas as células <td>/<th>, sem levar em conta <tr>.

    Usado para ler os rótulos das colunas: no arquivo exportado eles ficam
    soltos logo após a abertura de <table>, sem um <tr> envolvendo-os.
    """

    def __init__(self):
        super().__init__()
        self.cells = []
        self._current_cell = ''
        self._in_cell = False

    def handle_starttag(self, tag, attrs):
        if tag in ('td', 'th'):
            self._in_cell = True
            self._current_cell = ''

    def handle_endtag(self, tag):
        if tag in ('td', 'th'):
            self.cells.append(self._current_cell.strip())
            self._in_cell = False

    def handle_data(self, data):
        if self._in_cell:
            self._current_cell += data


def parse_header(content):
    first_tr = content.lower().find('<tr')
    header_html = content if first_tr == -1 else content[:first_tr]
    parser = _CellCollector()
    parser.feed(header_html)
    return parser.cells


def parse_input(filepath):
    with open(filepath, encoding='utf-8') as f:
        content = f.read()

    if parse_header(content) != EXPECTED_HEADER:
        raise InputFormatError(
            'Arquivo de entrada desatualizado. Favor gerar uma nova extração para corrigir.'
        )

    parser = TableParser()
    parser.feed(content)
    return parser.rows


def get_col(row, index):
    return row[index].strip() if index < len(row) else ''


def format_date(value):
    try:
        return datetime.strptime(value, '%Y-%m-%d').strftime('%d/%m/%Y')
    except ValueError:
        return value


def unique_output_path(output_dir, cpf):
    """Evita sobrescrever quando há mais de um formulário para o mesmo CPF:
    {cpf}.xlsx, depois {cpf}-2.xlsx, {cpf}-3.xlsx, etc."""
    path = os.path.join(output_dir, f'{cpf}.xlsx')
    i = 2
    while os.path.exists(path):
        path = os.path.join(output_dir, f'{cpf}-{i}.xlsx')
        i += 1
    return path


def insert_extra_segment_rows(ws, insert_at, amount):
    """Insere `amount` linhas extras em `insert_at` para acomodar mais de 4
    trechos, replicando a formatação da última linha de trecho do template
    (a linha logo acima do ponto de inserção) e corrigindo o que
    ws.insert_rows não ajusta sozinho: altura das linhas e mesclagens de
    células abaixo do ponto de inserção.
    """
    old_max_row = ws.max_row
    model_row = insert_at - 1

    # Precisa desfazer as mesclagens ANTES do insert_rows: o openpyxl desloca
    # as células físicas junto com o conteúdo, mas não atualiza os limites
    # guardados em merged_cells.ranges, então desfazer depois do insert_rows
    # tenta remover células de mesclagem que já não estão mais onde o
    # metadado (não deslocado) diz que deveriam estar.
    shifted_merges = []
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.min_row >= insert_at:
            shifted_merges.append((
                merged_range.min_row + amount, merged_range.min_col,
                merged_range.max_row + amount, merged_range.max_col,
            ))
            ws.unmerge_cells(str(merged_range))

    ws.insert_rows(insert_at, amount)

    for r in range(old_max_row, insert_at - 1, -1):
        ws.row_dimensions[r + amount].height = ws.row_dimensions[r].height

    for min_row, min_col, max_row, max_col in shifted_merges:
        ws.merge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)

    for i in range(amount):
        new_row = insert_at + i
        for col in range(1, ws.max_column + 1):
            ws.cell(row=new_row, column=col)._style = copy(ws.cell(row=model_row, column=col)._style)
        ws.row_dimensions[new_row].height = ws.row_dimensions[model_row].height


def generate_form(row, output_dir):
    cpf = get_col(row, 0).strip('"').strip()

    wb = openpyxl.load_workbook(get_resource_path(TEMPLATE_FILENAME))
    ws = wb.active

    ws['C4'] = get_col(row, 2)            # Identificação do evento
    ws['C5'] = get_col(row, 3)            # Data
    ws['C6'] = get_col(row, 4)            # Local

    servicos = [s.strip() for s in get_col(row, 5).split(',')]
    ws['C7'] = 'Passagens :  (X)' if 'Passagem aérea'       in servicos else 'Passagens :  ( )'
    ws['D7'] = 'Diárias: (X)'     if 'Diárias'              in servicos else 'Diárias: ( )'
    ws['E7'] = 'Terrestre: (X)'   if 'Transporte terrestre' in servicos else 'Terrestre: ( )'
    ws['F7'] = 'Aluguel de carro: (X)' if 'Aluguel de veículo' in servicos else 'Aluguel de carro: ( )'

    ws['B15'] = get_col(row, 6)             # Nome completo
    ws['C15'] = format_date(get_col(row, 7)) # Data de nascimento
    ws['D15'] = get_col(row, 9)             # Cargo/Função
    ws['E15'] = get_col(row, 8)             # CPF
    ws['F15'] = get_col(row, 11)            # Nome banco
    ws['G15'] = get_col(row, 12)            # Agência
    ws['H15'] = get_col(row, 13)            # DV agência
    ws['I15'] = get_col(row, 14)            # Conta corrente
    ws['J15'] = get_col(row, 15)            # DV conta
    ws['K15'] = get_col(row, 16)            # Poupança

    # Trechos: cada slot ocupa 8 colunas a partir de 19, na ordem: Tipo de Trecho
    # (Ida/Intermediário/Volta), Origem, Destino, Tipo Localidade de Destino,
    # Data, Período, Tipo Deslocamento, Observações sobre o deslocamento. Até
    # 10 slots reservados; para de ler no primeiro slot sem origem preenchida.
    #
    # Em algumas respostas (ex.: o 6º slot sempre, e eventualmente outros),
    # Tipo de Trecho e Origem vêm fisicamente trocados de posição — resquício
    # de reordenação de perguntas no Google Forms ao longo do tempo. Por isso
    # os dois primeiros campos são identificados pelo conteúdo (valor de
    # direção conhecido) em vez de por posição fixa.
    TRECHO_START = 19
    TRECHO_WIDTH = 8
    MAX_TRECHOS  = 10
    OFFSET_DESTINO, OFFSET_DATA, OFFSET_TURNO, OFFSET_TIPO_DESLOCAMENTO = 2, 4, 5, 6
    DIRECOES_VALIDAS = ('Ida', 'Intermediário', 'Volta')

    trechos = []
    for i in range(MAX_TRECHOS):
        base = TRECHO_START + i * TRECHO_WIDTH
        campo0 = get_col(row, base)
        campo1 = get_col(row, base + 1)
        if campo1 in DIRECOES_VALIDAS:
            direcao, origem = campo1, campo0
        else:
            direcao, origem = campo0, campo1
        if not origem:
            break
        trechos.append({
            'direcao':          direcao,
            'origem':           origem,
            'destino':          get_col(row, base + OFFSET_DESTINO),
            'data':             get_col(row, base + OFFSET_DATA),
            'turno':            get_col(row, base + OFFSET_TURNO),
            'tipo_deslocamento': get_col(row, base + OFFSET_TIPO_DESLOCAMENTO),
        })

    # Cada trecho vira sua própria linha: "Ida"/"Intermediário" preenchem
    # Data/Horário de ida (Data Volta/Horário de volta ficam em branco);
    # "Volta" preenche Data/Horário de volta (Data Ida/Horário de ida ficam
    # em branco). O Horário sempre combina Período + Tipo Deslocamento.
    output_rows = []
    for t in trechos:
        horario = f"{t['turno']} - {t['tipo_deslocamento']}"
        if t['direcao'] == 'Volta':
            output_rows.append({
                'origem': t['origem'], 'destino': t['destino'],
                'data_ida': '', 'hora_ida': '',
                'data_volta': t['data'], 'hora_volta': horario,
            })
        else:  # 'Ida' ou 'Intermediário'
            output_rows.append({
                'origem': t['origem'], 'destino': t['destino'],
                'data_ida': t['data'], 'hora_ida': horario,
                'data_volta': '', 'hora_volta': '',
            })

    SEGMENT_ROWS = [15, 16, 17, 18]
    extra_needed = len(output_rows) - len(SEGMENT_ROWS)
    if extra_needed > 0:
        insert_at = SEGMENT_ROWS[-1] + 1
        insert_extra_segment_rows(ws, insert_at, extra_needed)
        SEGMENT_ROWS = SEGMENT_ROWS + list(range(insert_at, insert_at + extra_needed))

    for out_row, seg in zip(SEGMENT_ROWS, output_rows):
        ws[f'L{out_row}'] = seg['origem']
        ws[f'M{out_row}'] = seg['destino']
        ws[f'N{out_row}'] = format_date(seg['data_ida'])
        ws[f'O{out_row}'] = seg['hora_ida']
        ws[f'P{out_row}'] = format_date(seg['data_volta'])
        ws[f'Q{out_row}'] = seg['hora_volta']

    justificativa_row = 19 + max(extra_needed, 0)
    ws[f'C{justificativa_row}'] = get_col(row, 99)      # Justificativa fora do prazo
    ws[f'C{justificativa_row + 1}'] = get_col(row, 100) # Observações

    output_path = unique_output_path(output_dir, cpf)
    wb.save(output_path)
    return output_path


def run_generation(input_file, base_output_dir, progress_callback=None, selected_rows=None):
    """
    Lê input_file e gera um xlsx por registro em base_output_dir/YYYYMMDD_HHMM/.
    progress_callback(msg) é chamado a cada evento de progresso.
    selected_rows, se fornecido, é usado em vez de parse_input(input_file).
    Retorna o caminho da pasta de saída criada.
    """
    run_stamp = datetime.now().strftime('%Y%m%d_%H%M')
    output_dir = os.path.join(base_output_dir, run_stamp)
    os.makedirs(output_dir, exist_ok=True)

    rows = selected_rows if selected_rows is not None else parse_input(input_file)

    if progress_callback:
        progress_callback(f'{len(rows)} registro(s) encontrado(s) em {os.path.basename(input_file)}')

    for i, row in enumerate(rows, 1):
        output_path = generate_form(row, output_dir)
        cpf = get_col(row, 0).strip('"')
        if progress_callback:
            progress_callback(f'  [{i}] CPF {cpf} → {os.path.basename(output_path)}')

    return output_dir


def main():
    print(f'Versão {get_version()}')

    if len(sys.argv) < 2:
        print(f'Uso: python {os.path.basename(sys.argv[0])} <arquivo_entrada.xls>')
        sys.exit(1)

    input_file = sys.argv[1]

    if not os.path.isfile(input_file):
        print(f'Erro: arquivo não encontrado: {input_file}')
        sys.exit(1)

    template_path = get_resource_path(TEMPLATE_FILENAME)
    if not os.path.isfile(template_path):
        print(f'Erro: template não encontrado: {template_path}')
        sys.exit(1)

    script_dir = os.path.dirname(os.path.abspath(__file__))
    try:
        run_generation(input_file, script_dir, progress_callback=print)
    except InputFormatError as exc:
        print(f'Erro: {exc}')
        sys.exit(1)
    print('Concluído.')


if __name__ == '__main__':
    main()
